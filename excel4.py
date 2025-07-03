import openpyxl
from openpyxl import Workbook
from openpyxl import Workbook , load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import json
import os 
import requests
# import streamlit as st
from openpyxl.drawing.image import Image


HANUAI_LOGO_PATH = "images/HanuAI.png"
ROADATHENA_LOGO_PATH = "images/RA-logo-1.png"

def process_json_data3(output_json_path, output_folder,roadId):

    road_id = roadId
    


    def fetch_road_data(roadId):
        api_url = f"https://ndd.roadathena.com/api/surveys/roads/{roadId}"
        response = requests.get(api_url,headers = { "Security-Password": "admin@123" })
        
        if response.status_code == 200:
            road_data = response.json()
            return road_data
        else:
            print(f"Failed to fetch data from API. Status code: {response.status_code}")
        
    
    def apply_common_formatting(sheet, survey_data):

        thick_border = Border(left=Side(style='medium'),
                              right=Side(style='medium'),
                              top=Side(style='medium'),
                              bottom=Side(style='medium'))

        # A1 - Title
        # sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nMC: {survey_data['mc']['name']}\nSubdivision: {survey_data['sub_division']['sub_division']}"
        sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {road_data['road']['ho']['name']}\nPIU : {road_data['road']['ro']['name']}"
        sheet['A1'].font = Font(b=True, size=14)
        sheet['A1'].alignment = Alignment(wrap_text=True, horizontal='center')
        sheet.merge_cells('A1:M1')  
        sheet['A1'].border = thick_border

        for col in range(1, 14):  # Columns A to L
            cell = sheet.cell(row=1, column=col)
            cell.border = thick_border


        # A2 - Date of survey
        sheet["A2"] = f"Date of survey: {survey_data['survey_date']}"
        sheet['A2'].font = Font(size=13)
        sheet.merge_cells('A2:K2') 
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet['A2'].border = thick_border

        for col in range(1, 12):  # Columns A to L
            cell = sheet.cell(row=2, column=col)
            cell.border = thick_border

        # A3 - Segment name
        sheet["A3"] = f"Segment name: {survey_data['road_name']}"
        sheet.merge_cells('A3:K3')  # Merge cells A3 to E3
        sheet['A3'].font = Font(size=13, b=True)
        sheet['A3'].alignment = Alignment(horizontal='center')
        sheet['A3'].border = thick_border

        for col in range(1, 12):  # Columns A to L
            cell = sheet.cell(row=3,column=col)
            cell.border = thick_border

        # A4 - Start Chainage
        sheet["A4"] = f"Start Chainage"
        sheet['A4'].font = Font(size=13, b=True)
        sheet['A4'].alignment = Alignment(horizontal='center')
        sheet['A4'].border = thick_border

        # B4 - Start Chainage value
        sheet["B4"] = f"{survey_data['start_chainage']}"
        sheet['B4'].font = Font(size=13)
        sheet['B4'].alignment = Alignment(horizontal='center')
        sheet['B4'].border = thick_border

        # A5 - End Chainage
        sheet["A5"] = f"End Chainage"
        sheet['A5'].font = Font(size=13, b=True)
        sheet['A5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet['A5'].border = thick_border

        # B5 - End Chainage value
        sheet["B5"] = f"{survey_data['end_chainage']}"
        sheet['B5'].font = Font(size=13)
        sheet['B5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet['B5'].border = thick_border

        img = Image(ROADATHENA_LOGO_PATH)
        img.anchor = 'A1'
        sheet.add_image(img)

        img = Image(HANUAI_LOGO_PATH)
        img.anchor = 'K1'
        sheet.add_image(img)

        sheet.row_dimensions[1].height = 60
        sheet.row_dimensions[2].height = 30
        sheet.row_dimensions[3].height = 30
        sheet.row_dimensions[4].height = 50
        sheet.row_dimensions[5].height = 50
        sheet.row_dimensions[6].height = 50

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20

        return sheet

    def create_detailed_report(data, survey_data, output_folder):
        wb = load_workbook(os.path.join(output_folder, f"{roadId}_formatted.xlsx"))

        # Check and create "Encroachment Signs" sheet
        encroachment_sheet_name = "Encroachment Signs"
        if encroachment_sheet_name in wb.sheetnames:
            del wb[encroachment_sheet_name]
        ws_encroachment = wb.create_sheet(encroachment_sheet_name)
        ws_encroachment.title = encroachment_sheet_name
        
    
        
        # Formatting for Encroachment Signs sheet
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        ws_encroachment["A6"] = "Enchrochment Number"
        ws_encroachment["B6"] = "Timestamp on processed video"
        ws_encroachment["C6"] = "Asset type"
        ws_encroachment["D6"] = "Side"
        ws_encroachment["E6"] = "Category"
        ws_encroachment["F6"] = "Latitude"
        ws_encroachment["G6"] = "Longitude"
        ws_encroachment["H6"] = "Distance from start point in meters"
        ws_encroachment["I6"] = "Image Link"

        for col in range(1, 10):
            cell = ws_encroachment.cell(row=6, column=col)
            cell.font = Font(b=True, size=12)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = PatternFill(start_color='c0d3eb', end_color='c0d3eb', patternType='solid')

        # Filtered data
        filtered_data = [
           item for item in (data.get("assets", []) + data.get("anomalies", []))
        if (
           item.get("Asset type") in ["ADVERTISEMENT_ENCHROACHMENT_SIGNS", "NON_STANDARD_INFOMATORY_SIGNS" ]
           or item.get("Anomaly type") in ["ADVERTISEMENT_ENCHROACHMENT_SIGNS", "NON_STANDARD_INFOMATORY_SIGNS" ]
           )
        ]    


        # Populate the "Encroachment Signs" sheet
        serial_number = 1
        for i, row_item in enumerate(filtered_data, start=7):
            ws_encroachment[f'A{i}'] = serial_number
            ws_encroachment[f'B{i}'] = row_item["Timestamp on processed video"]
            ws_encroachment[f'C{i}'] = row_item.get("Asset type") or row_item.get("Anomaly type")
            ws_encroachment[f'D{i}'] = row_item["Side"]
            ws_encroachment[f'E{i}'] = row_item["category"]
            ws_encroachment[f'F{i}'] = row_item["Latitude"]
            ws_encroachment[f'G{i}'] = row_item["Longitude"]
            ws_encroachment[f'H{i}'] = row_item["Distance"]
            ws_encroachment[f'I{i}'] = row_item["image"]

            for col in range(1, 10):
                cell = ws_encroachment.cell(row=i, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            serial_number += 1

        apply_common_formatting(ws_encroachment, survey_data)

            

        # Save workbook
        output_file_path = os.path.join(output_folder, f"{roadId}_formatted.xlsx")
        wb.save(output_file_path)
        print(f"Encroachment Signs sheet added to {output_file_path}.")

    # Load JSON data
    file_path = os.path.join(output_json_path, f"road_{roadId}.json")

    with open(file_path, "r") as file:
        json_data = json.load(file)

    road_data = fetch_road_data(road_id)
    
    if road_data:
        try:
            seg = road_data["road"]["name"]
            start_chainage = road_data["road"]["start_chainage"]
            end_chainage = road_data["road"]["end_chainage"]
            mc_name = road_data["road"]["assigned_to"]["username"]
            sub_division_name = road_data["road"]["assigned_to"]["sub_division"]
            key = road_data["created_at"].split("T")[0]
        except KeyError as e:
            print(f"KeyError: {e}")
            return

        survey_data = {
            "road_name": seg,
            "start_chainage": start_chainage,
            "end_chainage": end_chainage,
            "mc": {"name": mc_name},
            "sub_division": {"sub_division": sub_division_name},
            "survey_date": key
        }

        # Create detailed report
        # print(json_data[0])
        create_detailed_report(json_data, survey_data, output_folder)
    else:
        print("Failed to fetch road data. Exiting...")

# Example usage
# output_folder = "C:/Users/manav/OneDrive/Desktop/furniture-v4.0"
# output_json_path = "C:/Users/manav/OneDrive/Desktop/furniture-v4.0/final.json"
# process_json_data(output_json_path, output_folder)
     
