import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

def reformat_excel_for_a4(input_file, output_file):
    # Load the Excel file
    wb = openpyxl.load_workbook(input_file)

    # Iterate through each sheet in the workbook
    for sheet in wb.worksheets:
        # Set the page setup to A4 size for all sheets
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        total_columns = sheet.max_column

        # Check if the sheet is in landscape orientation
        if sheet.title != "Pavement Anomaly":
            # Set landscape orientation
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            # Apply margins for landscape mode (top = 0.4)
            sheet.page_margins = PageMargins(left=0.15, right=0, top=0.25, bottom=0.25)

            # Adjust column widths to fit the page (for landscape mode)
            total_columns = sheet.max_column
            available_width = 14.5  # Available width for landscape mode
            width_per_column = available_width / total_columns

            for col in range(1, total_columns + 1):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = width_per_column * 10  # Adjust column widths

            # Apply text wrapping, font size, and alignment for landscape sheets
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:  # Ensure the cell has content before changing formatting
                        # Enable text wrapping
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.font = Font(size=8)  # Adjust font size to 10 to fit more content

            # Dynamically adjust row heights to fit wrapped text
            # Set row height for the first five rows
            
            sheet.row_dimensions[1].height = 60 
            for row_idx in range(2, 6):  # Rows 1 to 5
                sheet.row_dimensions[row_idx].height = 30 # Auto-adjust row height for wrapped text
        else:
            # For portrait sheets (e.g., "Pavement Anomaly"), set portrait orientation
            sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
            # Set margins for portrait mode
            sheet.page_margins = PageMargins(left=0.25 , right=0 , top=0.25, bottom=0.25)
            
            available_width = 10
            width_per_column = available_width / total_columns
        
            for col in range(1, total_columns + 1):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = width_per_column * 10  # Adjusting column widths

            # Reduce the font size for all cells
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:  # Ensure the cell has content before changing the font
                        cell.font = Font(size=7)  # Set font size to 8 to fit more content
            
            
    # Save the modified file
    wb.save(output_file)
    print(f"Reformatted Excel file saved as {output_file}") 
    
    
# Usage example
import os

## input excels path 
input_folder = "op" 

## op excels path 
op_folder = "pdfs_reformatted_a4"
if not os.path.exists(op_folder):
    os.makedirs(op_folder)
    

## excel fiels in foldree 
for file in os.listdir(input_folder):
    if file.endswith(".xlsx"):
        
        input_file = os.path.join(input_folder, file)
        file_name = file
        # input_file = f'excels/{file_name}'

        output_file = f'{op_folder}/{file_name}'

        print(input_file)
        print(output_file)

        reformat_excel_for_a4(input_file, output_file)

