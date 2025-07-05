import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import json 
import os
import id
# from furniture.code.extra_utils import *
from openpyxl import Workbook, load_workbook
from collections import defaultdict
import math 
def get_max_distance(gpx_data):
    max_value = max((x['distanceInMeters'] for x in gpx_data.values()), default=0) 
    # max_value = ((max_value + 499) // 500) * 500
    return math.ceil(max_value)




    
## get closest timestamp 
def find_closest_timestamp(target_distance, timestamp_dict , max_distance_differnce = 500  ):
    min_difference = float('inf')
    closest_values = None
    for values in timestamp_dict.values():
        distance = values['distanceInMeters']
        if abs(distance - float(target_distance)) <= max_distance_differnce:  # Check if the distance is within 100 meters
            difference = abs(distance - float(target_distance))
            if difference < min_difference:
                min_difference = difference
                closest_values = values
    return closest_values



def generate_counts_dict():
    """Generate the default structure for counts."""
    return {
        "CHEVRON": 0,
        "CAUTIONARY_WARNING_SIGNS": 0,
        "HAZARD": 0,
        "PROHIBITORY_MANDATORY_SIGNS": 0,
        "INFORMATORY_SIGNS": 0,
    }


def process_item(counts, item, side_key):
    """Process an item and update the counts."""
    if side_key not in counts:
        counts[side_key] = {}

    # Determine the type key (Asset or Anomaly)
    key = item.get("Asset type") or item.get("Anomaly type")
    if not key:
        return counts  # nothing to process

    # Initialize if not already in the dictionary
    if key not in counts[side_key]:
        counts[side_key][key] = 0

    counts[side_key][key] += 1
    return counts



def add_gps_data(result, key, gpx_data, distance_start, distance_end):
    """Add GPS data to the result for the given range."""
    closest_from_values = find_closest_timestamp(distance_start, gpx_data)
    closest_to_values = find_closest_timestamp(distance_end, gpx_data)

    if closest_from_values is not None:
        result[key]['lat'] = closest_from_values.get('lat', 0)
        result[key]['lng'] = closest_from_values.get('lng', 0)

    if closest_to_values is not None:
        result[key]['to_lat'] = closest_to_values.get('lat', 0)
        result[key]['to_lng'] = closest_to_values.get('lng', 0)
        
        
## code of the count data count data of detected anomalies 
def parse_json(json_data, max_distance, gpx_data, road_name, first_gap  , interval=500):
    result = {}
    distance_ranges = []
    start = 0

    # Generate distance ranges
   # max_distance = 47000

    while start < max_distance:
        print("yes")
        end = first_gap if start == 0 else min(start + interval, max_distance)
        distance_ranges.append((start, end))
        start = end
        print("start end" , start , end )
    print("no")
    print(distance_ranges, "printing ... distance ranges ", max_distance, "max distances")

    # Process each distance range
    for i, distance_range in enumerate(distance_ranges):
        

        if "SR" in road_name or "SL" in road_name:
         counts = {
        "Left": defaultdict(int), 
        "Right": defaultdict(int),
        }
        else:
         counts = {
        "Avenue": defaultdict(int),
        "Median": defaultdict(int),
        "Center": defaultdict(int),
    }


           
        # Default values for the range
        range_key = distance_range[0]
        result[range_key] = counts 
        
        for item in json_data['assets'] + json_data['anomalies']:
        # Extract distance
            try:
                distance = float(item['Distance'])
            except (ValueError, KeyError):
                continue  # 

            # Check if distance falls in the current range
            if distance_range[0] <= distance < distance_range[1]:
                # Determine the side
                side = item.get('Side', 'Center').title()
                counts = process_item(counts, item, side)

                # Add counts and GPS data to the result
            result[range_key] = counts
            add_gps_data(result, range_key, gpx_data, distance_range[0], distance_range[1])
            

    return result

def previous_value_divisible_by_500(value):
    if value < 500:
        return 0
    return (value // 500) * 500


def previous_next_divisible_by_500(value):
    
    previous_value = (value // 500) * 500
    next_value = previous_value + 500 
    return next_value


# Function to handle sorting of the assets based on distance
def sort_assets_by_distance(data):
    # Sort the assets, placing None (null) distances at the end of the list
    sorted_assets = sorted(
        data['assets'], 
        key=lambda asset: (asset['Distance'] is None, asset['Distance'])
    )
    return sorted_assets


def process_json_data2(output_json_path, gpx_json_path , output_folder, roadId):
    
    def fetch_road_data(roadId):
        api_url = f"https://ndd.roadathena.com/api/surveys/roads/{roadId}"
        response = requests.get(api_url, headers = { "Security-Password": "admin@123" })
        
        if response.status_code == 200:
            json_response = response.json()
            lhr_side = json_response['road']['LHR_side']
            rhr_side = json_response['road']['RHR_side']

            print(f"LHR_side: {lhr_side}, RHR_side: {rhr_side}")
            return response.json()
        else:
            print(f"Failed to fetch data from API. Status code: {response.status_code}")
            return None

    def apply_common_chainage_formatting(sheet, survey_data):
        thick_border = Border(left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))
        
        # A1 - Title
        sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {road_data['road']['ho']['name']}\nPIU : {road_data['road']['ro']['name']}"
        sheet['A1'].font = Font(b=True, size=14)
        sheet['A1'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.merge_cells('A1:L1')  # Merge cells A1 to L1

        # Apply border to the merged range A1:L1
        for col in range(1, 13):  # Columns A to L
            cell = sheet.cell(row=1, column=col)
            cell.border = thick_border


        # A2 - Date of survey
        sheet["A2"] = f"Date of survey: {survey_data['survey_date']}"
        sheet['A2'].font = Font(b=True,size=13)
        sheet.merge_cells('A2:G2')
        sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A2'].border = thick_border

        for col in range(1, 8):  # Columns A to L
            cell = sheet.cell(row=2, column=col)
            cell.border = thick_border


        # A3 - Segment name
        sheet["A3"] = f"Segment name: {survey_data['road_name']}"
        sheet.merge_cells('A3:G3')
        sheet['A3'].font = Font(size=13, b=True)
        sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A3'].border = thick_border
        

        for col in range(1, 8):  # Columns A to L
            cell = sheet.cell(row=3, column=col)
            cell.border = thick_border


        # A4 - Start Chainage
        sheet["A4"] = f"Start Chainage"
        sheet.merge_cells('A4:B4')
        sheet['A4'].font = Font(size=13, b=True)
        sheet['A4'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A4'].border = thick_border
        
        for col in range(1, 4):  # Columns A to L
            cell = sheet.cell(row=4, column=col)
            cell.border = thick_border


        # B4 - Start Chainage value
        sheet["C4"] = f"{survey_data['start_chainage']}"
        sheet['C4'].font = Font(size=13)
        sheet.merge_cells('C4:D4')
        sheet['C4'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['C4'].border = thick_border
        for col in range(1, 4): 
            cell = sheet.cell(row=4, column=3)
            cell.border = thick_border


        # A5 - End Chainage
        sheet["A5"] = f"End Chainage"
        sheet.merge_cells('A5:B5')
        sheet['A5'].font = Font(size=13, b=True)
        sheet['A5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet['A5'].border = thick_border
        
        for col in range(1, 4): 
            cell = sheet.cell(row=5, column=col)
            cell.border = thick_border
            
        
        end_chainage_value = survey_data.get('end_chainage', 'N/A') 
        sheet["C5"] = f"{end_chainage_value}"
        sheet['C5'].font = Font(size=13)
        sheet.merge_cells('C5:D5')
        sheet['C5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet['C5'].border = thick_border
        for col in range(1, 4): 
            cell = sheet.cell(row=5, column=3)
            cell.border = thick_border

        print(f"End Chainage Value: {end_chainage_value}")

        # Adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.row_dimensions[1].height = 80
        sheet.row_dimensions[2].height = 30
        sheet.row_dimensions[3].height = 30
        sheet.row_dimensions[4].height = 50
        sheet.row_dimensions[5].height = 50
        sheet.row_dimensions[6].height = 50

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20

        return sheet
    

    def countingExcelReport(survey_data):

        # wb = Workbook()
        # wb = load_workbook(os.path.join(output_folder, f"{road_data['road']['name']}_formatted.xlsx"))
        wb = load_workbook(os.path.join(output_folder, f"{roadId}_formatted.xlsx"))
        
        # report_name = 'Road Furniture Chainage wise report'
        report_name = 'Furniture Chainage report'
        if report_name  in wb.sheetnames:
            
            del wb[report_name]
                
        ws = wb.create_sheet(report_name)
        ws.title = report_name
        wb.active = ws
        ws = apply_common_chainage_formatting(ws, survey_data)
        # print(ws)
        

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        if "SR" or "SL" in road_name:
            headers = {
                "A7": "Chainage", "A8": "From", "B8": "To",
                "C8": "From", "E8": "To", "C7": "Geo-Location(Lat,Long)",
                "G7": "Survey Name", "H7": "Survey Date", "I7": "Direction",
                "J7": "Side", "K7": "Furniture Assets", "K8": "CHEVRON", "K10": "Left", "L10": "Right",
                "M8": "CAUTIONARY_WARNING_SIGNS", "M10": "Left", "N10": "Right",
                "O8": "HAZARD", "O10": "Left", "P10": "Right", 
                "Q8": "PROHIBITORY_MANDATORY_SIGNS", "Q10": "Left", "R10": "Right",
                "S8": "INFORMATORY_SIGNS", "S10": "Left", "T10": "Right", "A6": report_name,
            }
        else:
            headers = {
                "A7": "Chainage", "A8": "From", "B8": "To",
                "C8": "From", "E8": "To", "C7": "Geo-Location(Lat,Long)",
                "G7": "Survey Name", "H7": "Survey Date", "I7": "Direction",
                "J7": "Side", "K7": "Furniture Assets", "K8": "CHEVRON","K10":"Avenue","L10":"Median",
                "M8": "CAUTIONARY_WARNING_SIGNS","M10":"Avenue","N10":"Median", "O8": "HAZARD","O10":"Avenue","P10":"Median", "Q8": "PROHIBITORY_MANDATORY_SIGNS",
                "Q10":"Avenue","R10":"Median","S8": "INFORMATORY_SIGNS", "S10":"Avenue","T10":"Median","A6": report_name,
            }

        for cell, value in headers.items():
            ws[cell] = value

        merge_ranges = [
            ("A6", "O6"), ("A7", "B7"), ("C7", "F7"),
            ("G7", "G9"), ("H7", "H9"), ("I7", "I9"), ("J7", "J9"),
            ("K7", "T7"), ("A8", "A9"), ("B8", "B9"), ("C8", "D9"),
            ("E8", "F9"), ("K8", "L9") ,("M8", "N9"),("O8","P9"),('Q8','R9'),('S8','T9'),('U8','V9')
        ]

        COLOR = "6AC9FF"
        for start, end in merge_ranges:
            ws.merge_cells(f"{start}:{end}")

        for row in range(6, 11):
            for col in "ABCDEFGHIJKLMNOPQRST":
                cell = ws[f'{col}{row}']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name="Arial", bold=True)
                cell.fill = PatternFill(start_color=COLOR, end_color=COLOR, fill_type="solid")
                cell.border = thin_border

        column_widths = {
            'A': 15, 'B': 15, 'C': 16, 'D': 16, 'E': 16,
            'F': 16, 'G': 25, 'H': 15, 'I': 25, 'J': 12,
            'K': 12, 'L': 12, 'M': 16, 'N': 12, 'O': 16 , 'P': 16,'Q': 16,'R': 16,'S': 16,'T': 16 
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        row_heights = {1: 55, 2: 40, 3: 40, 4: 40}

        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        return wb


    def chainageWiseCounting(survey_data, json_data , side , value_diff ):

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        print(survey_data.get("side") , "getting...")
        # side = survey_data.get("side")
        # side = "RHS"

        if side == "LHS":
            survey_direction = "Increasing"
        elif side == "RHS":
            survey_direction = "Decreasing"
        else:
            survey_direction = "None"            
        print("fux" , side )
    
        if survey_data["start_chainage"]:
            
            # start_chainage = int(survey_data["start_chainage"].split("+")[0]) * 1000 + int(survey_data["start_chainage"].split("+")[1])
            start_chainage = int(survey_data["start_chainage"]) 
        
        else:
            start_chainage = 0 
            
            
        survey_date = survey_data["survey_date"]
        survey_name = survey_data["survey_name"]
        survey_direction = survey_direction
        survey_side = {side}
        print(survey_side)

        wb = countingExcelReport(survey_data)
        report_name = 'Furniture Chainage report'
        ws = wb[report_name]
        wb.active = ws 

        start_row = 11
        to_chainage = 0 

        end_chainage_value = survey_data.get('end_chainage', None)
        
        chainage_diff = value_diff
            
        # if chainage_diff > 0:
        #     side = "RHS"
        # else :
        #     side = "LHS"
        
        # start_gap = start_chainage
        
        if end_chainage_value is None:
            print("End chainage value is not defined. Exiting the process.")
        else:
            try:
                end_chainage_value = int(end_chainage_value)
                print(end_chainage_value)
            except ValueError:
                print(f"Invalid end chainage value: {end_chainage_value}. It must be a valid integer.")
                exit()
                    
        print("chainage ...value" , side )
        for i, (key, data) in enumerate(json_data.items()):
            
            current_row = start_row + i
            
            if i == 0 :
                
                if side == "RHS":
                
                    from_chainage = int(start_chainage)
                    
                    if start_chainage != 0 and int(start_chainage)%500 == 0 :
                        to_chainage = from_chainage - 500
                    
                    if start_chainage != 0 and int(start_chainage)%500 != 0 :
                        print("true")
                        
                        to_chainage = previous_value_divisible_by_500(int(start_chainage))
                        
                if side == "LHS":
                    
                    from_chainage = int(start_chainage)

                    
                    if start_chainage != 0 and int(start_chainage)%500 == 0 :
                        to_chainage = from_chainage + 500
                    
                    if start_chainage !=0 and int(start_chainage)%500 != 0 :
                        
                        to_chainage = previous_next_divisible_by_500(int(start_chainage))
       
            else:
                
                if from_chainage == start_chainage and side =="RHS":
                    
                    from_chainage = to_chainage
                    to_chainage = from_chainage - 500
                    
                elif side == "RHS":
                    
                    from_chainage = from_chainage - 500
                    to_chainage = from_chainage - 500
                    
                else: 
                    
                    from_chainage = to_chainage 
                    to_chainage = to_chainage + 500


                # Check if to_chainage exceeds the chainage_limit
                if side == "LHS":
                    
                    if end_chainage_value :
                        if to_chainage > end_chainage_value and to_chainage - end_chainage_value < 500:
                            to_chainage = end_chainage_value
                        
                    
                        if to_chainage > end_chainage_value:
                            print(f"Chainage limit exceeded: {end_chainage_value}. Stopping further processing.")
                            break
                    
                else:
                    if side =="RHS":
                        if to_chainage < 0:
                            print(f"Chainage limit exceeded: {end_chainage_value}. Stopping further processing.")
                            break 
                        
                        if i == len(json_data) -1 :
                            to_chainage = end_chainage_value
                        
                        
                            
                    else:
                        
                        if to_chainage == end_chainage_value:
                            print(f"Chainage limit exceeded: {end_chainage_value}. Stopping further processing.")
                            break
                        

            print(start_chainage , from_chainage, to_chainage) 
            
            
            try:
                if i == 0:
                    data2 = json_data[(chainage_diff)]
                    ws[f'E{current_row}'] = data2["lat"]
                    ws[f'F{current_row}'] = data2["lng"]
                elif i == len(json_data) - 1:
                    data2 = json_data[(from_chainage)]
                    ws[f'E{current_row}'] = data2["lat"]
                    ws[f'F{current_row}'] = data2["lng"]
                
                else:
                    data2 = json_data[(500*i + chainage_diff)]
                    ws[f'E{current_row}'] = data2["lat"]
                    ws[f'F{current_row}'] = data2["lng"]
            except KeyError:
                print("End key error")
                

            if side == "LHS":
                ws[f'A{current_row}'] = from_chainage 
                ws[f'B{current_row}'] = to_chainage 
                ws[f'J{current_row}'] = side
                
                
            else:
                
                
                ws[f'A{current_row}'] =  from_chainage
                ws[f'B{current_row}'] =  to_chainage 
                ws[f'J{current_row}'] = side
        
            try:

                ws[f'C{current_row}'] = data.get("lat", 0)  # Default to 0 if "lat" doesn't exist or is None
                ws[f'D{current_row}'] = data.get("lng", 0)  # Default to 0 if "lng" doesn't exist or is None
                ws[f'E{current_row}'] = data.get("to_lat", 0)  # Default to 0 if "to_lat" doesn't exist or is None
                ws[f'F{current_row}'] = data.get("to_lng", 0)

            except KeyError as e:

                ws[f'C{current_row}'] = 0
                ws[f'D{current_row}'] = 0
                ws[f'E{current_row}'] = 0
                ws[f'F{current_row}'] = 0
            
            ws[f'G{current_row}'] = survey_name
            ws[f'H{current_row}'] = survey_date
            ws[f'I{current_row}'] = survey_direction

            def get_informatory_count(section_data):
                return (
                    section_data.get("INFORMATORY_SIGNS", 0) +
                    section_data.get("DIGITAL_SPEED_DISPLAY_SIGNS", 0) +
                    section_data.get("VARIABLE_MESSAGE_SIGNS", 0)
                )

            # if "SR" in road_name:

            # ws[f'K{current_row}'] = data["Left"].get("CHEVRON", 0)
            # ws[f'L{current_row}'] = data["Right"].get("CHEVRON", 0)

            # ws[f'M{current_row}'] = data["Left"].get("CAUTIONARY_WARNING_SIGNS", 0)
            # ws[f'N{current_row}'] = data["Right"].get("CAUTIONARY_WARNING_SIGNS", 0)

            # ws[f'O{current_row}'] = data["Left"].get("HAZARD", 0)
            # ws[f'P{current_row}'] = data["Right"].get("HAZARD", 0)

            # ws[f'Q{current_row}'] = data["Left"].get("PROHIBITORY_MANDATORY_SIGNS", 0)
            # ws[f'R{current_row}'] = data["Right"].get("PROHIBITORY_MANDATORY_SIGNS", 0)

            # ws[f'S{current_row}'] = get_informatory_count(data["Left"])
            # ws[f'T{current_row}'] = get_informatory_count(data["Right"])


            if id.mcw_id: 

                ws[f"K{current_row}"] = data["Avenue"].get("CHEVRON", 0)  
                ws[f"L{current_row}"] = data["Median"].get("CHEVRON", 0)  
                
                ws[f"M{current_row}"] = data["Avenue"].get("CAUTIONARY_WARNING_SIGNS", 0)  
                ws[f"N{current_row}"] = data["Median"].get("CAUTIONARY_WARNING_SIGNS", 0)  
                
                ws[f"O{current_row}"] = data["Avenue"].get("HAZARD", 0) 
                ws[f"P{current_row}"] = data["Median"].get("HAZARD", 0) 
                
                ws[f"Q{current_row}"] = data["Avenue"].get("PROHIBITORY_MANDATORY_SIGNS", 0)  
                ws[f"R{current_row}"] = data["Median"].get("PROHIBITORY_MANDATORY_SIGNS", 0)  
                
                ws[f"S{current_row}"] = get_informatory_count(data["Avenue"])  
                ws[f"T{current_row}"] = get_informatory_count(data["Median"])

            else:
            
                left_data = data.get("Left", {})
                right_data = data.get("Right", {})

                ws[f'K{current_row}'] = left_data.get("CHEVRON", 0)
                ws[f'L{current_row}'] = right_data.get("CHEVRON", 0)

                ws[f'M{current_row}'] = left_data.get("CAUTIONARY_WARNING_SIGNS", 0)
                ws[f'N{current_row}'] = right_data.get("CAUTIONARY_WARNING_SIGNS", 0)

                ws[f'O{current_row}'] = left_data.get("HAZARD", 0)
                ws[f'P{current_row}'] = right_data.get("HAZARD", 0)

                ws[f'Q{current_row}'] = left_data.get("PROHIBITORY_MANDATORY_SIGNS", 0)
                ws[f'R{current_row}'] = right_data.get("PROHIBITORY_MANDATORY_SIGNS", 0)

                ws[f'S{current_row}'] = get_informatory_count(left_data)
                ws[f'T{current_row}'] = get_informatory_count(right_data)


            FONT = "Microsoft Sans Serif"
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O','P','Q', 'R', 'S', 'T']:
                cell = ws[f'{col}{current_row}']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name=FONT, size=10, bold=False)
                cell.border = thin_border

        # output_file_path = os.path.join(output_folder, f"{survey_data['road_name']}_formatted.xlsx")
        output_file_path = os.path.join(output_folder, f"{roadId}_formatted.xlsx")
        wb.save(output_file_path)

    road_data = fetch_road_data(roadId)



    if road_data:
        try:
            road_name = road_data["road"]["name"]
            
            start_chainage = road_data["road"]["start_chainage"]
            end_chainage = road_data["road"]["end_chainage"]
            
            
            print(start_chainage , end_chainage , "starting ....") 
            
            if start_chainage is not None :
                
                start_chainage = float(start_chainage)
                end_chainage = float(end_chainage)
                
                print("indise")
                
            
                chainage_diff = end_chainage - start_chainage
                
                print(chainage_diff )
                
                
                if chainage_diff < 0:
                    
                    side = "RHS"
                    next_value = previous_value_divisible_by_500(int(start_chainage))
                    value_differnce =  int(start_chainage) - next_value
                    print("next value is " , next_value , value_differnce)
                    
                    if value_differnce == 0 :
                        value_differnce = 500 
                    
                    
                else:
                    side = "LHS"
                    next_value = previous_next_divisible_by_500(int(start_chainage))
                    value_differnce = next_value - int(start_chainage)
                    print("hello i am here ......")
                    
                    if value_differnce == 0:
                        value_differnce = 500 
            
            
            else:
                side = "LHS"
                chainage_diff = 0 
                next_value = 500 
                value_differnce = 500 
                
                print("outer condition ...")
                
                
                
            print(side , chainage_diff, next_value, value_differnce , "...........")
                
            
            
            

            mc_name = road_data["road"]["assigned_to"]["username"]
            survey_name = road_data["road"]["name"]
            survey_date = road_data["created_at"].split("T")[0]
            sub_division_name = road_data["road"]["assigned_to"]["sub_division"]
            # print(road_data["road"]["LHR_side"] , "checking ")
            
            
            # if  road_data["road"]["LHR_side"]:
            #     side = "LHS"
            # else:
            #     side = "RHS"

            print("side is" , side )
            
            
        except KeyError as e:
            print(f"KeyError: {e}")
            return

        survey_data = {
            "road_name": road_name,
            "start_chainage": start_chainage,
            "end_chainage": end_chainage,
            "end_chainage": end_chainage,
            "mc": {"name": mc_name},
            "sub_division": {"sub_division": sub_division_name},
            "survey_date": survey_date,
            "survey_name": survey_name,
            "side": side
        }
        
        
        

        GPX_JSON = os.path.join(gpx_json_path , f"gpx_road_{roadId}.json")
        with open(GPX_JSON, 'r') as f:
            gpx_data = json.load(f)
        
        DETECTION_JSON = os.path.join(output_json_path, f"road_{roadId}.json")
        with open(DETECTION_JSON, 'r') as f:
            detections_data = json.load(f)  
            
        detections_data['assets'] = sort_assets_by_distance(detections_data)
            
            
            
            
        

        
        
        max_distance = get_max_distance(gpx_data)
        # max_distance_ns = get_max_distance(gpx_data)

        # max_distance = previous_next_divisible_by_500(max_distance_ns)
        print("max distance is before parse json " , max_distance , value_differnce)

        # max_distance2 = survey_data.get('end_chainage', None)
        parsed_result = parse_json(detections_data,max_distance , gpx_data, road_name , value_differnce )
        
        
        print("after parse  json file.." , max_distance)
        
        # print("parsed results ..." , parsed_result )
        COUNT_JSON = os.path.join("count_jsons" ,f"count_{roadId}.json")

        
        with open(COUNT_JSON, 'w') as f:
            json.dump(parsed_result, f, indent=4)
            
            
        # print("surve" , survey_data["side"])
        chainageWiseCounting(survey_data, parsed_result , side , value_differnce )
