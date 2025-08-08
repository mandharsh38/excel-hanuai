import openpyxl
from openpyxl import Workbook
from openpyxl import Workbook , load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import json
import os 
import requests
from openpyxl.drawing.image import Image
from io import BytesIO

import sys

def fatal_error(msg):
    print(f"{msg}")
    sys.exit(1)


HANUAI_LOGO_PATH = "images/HanuAI.png"
ROADATHENA_LOGO_PATH = "images/RA-logo-1.png"

def process_json_data(output_json_path, output_folder,roadId):

    road_id = roadId

    def fetch_road_data(roadId):
        api_url = f"https://ndd.roadathena.com/api/surveys/roads/{roadId}"
        #api_url = f"https://dev.roadathena.com/api/surveys/roads/{roadId}"

        response = requests.get(api_url,headers = { "Security-Password": "admin@123" })
        
        if response.status_code == 200:
            road_data = response.json()
            return road_data
        else:
            fatal_error(f"Failed to fetch data from API. Status code: {response.status_code}")
        
    
    def apply_common_formatting(sheet, survey_data):

        thick_border = Border(left=Side(style='medium'),
                              right=Side(style='medium'),
                              top=Side(style='medium'),
                              bottom=Side(style='medium'))

        
        sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {road_data['road']['ho']['name']}\nPIU : {road_data['road']['ro']['name']}"
        # sheet["A1"] = f"AI Based road condition assessment detailed report by ROAD ATHENA\nRO : {'test'}\nPIU : {'test'}"
        sheet['A1'].font = Font(b=True, size=14)
        sheet['A1'].alignment = Alignment(wrap_text=True, horizontal='center')
        sheet.merge_cells('A1:M1')  
        sheet['A1'].border = thick_border

        for col in range(1, 14):  # Columns A to L
            cell = sheet.cell(row=1, column=col)
            cell.border = thick_border


        # A2 - Date of survey
        sheet["A2"] = f"Date of survey: {survey_data['survey_date']}"
        sheet['A2'].font = Font(size=13)
        sheet.merge_cells('A2:K2') 
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet['A2'].border = thick_border

        for col in range(1, 12):  # Columns A to L
            cell = sheet.cell(row=2, column=col)
            cell.border = thick_border

        # A3 - Segment name
        sheet["A3"] = f"Segment name: {survey_data['road_name']}"
        sheet.merge_cells('A3:K3')  # Merge cells A3 to E3
        sheet['A3'].font = Font(size=13, b=True)
        sheet['A3'].alignment = Alignment(horizontal='center')
        sheet['A3'].border = thick_border

        for col in range(1, 12):  # Columns A to L
            cell = sheet.cell(row=3,column=col)
            cell.border = thick_border

        sheet["A4"] = f"Start Chainage"
        sheet['A4'].font = Font(size=13, b=True)
        sheet['A4'].alignment = Alignment(horizontal='center')
        sheet['A4'].border = thick_border

        sheet["B4"] = f"{survey_data['start_chainage']}"
        sheet['B4'].font = Font(size=13)
        sheet['B4'].alignment = Alignment(horizontal='center')
        sheet['B4'].border = thick_border

        sheet["A5"] = f"End Chainage"
        sheet['A5'].font = Font(size=13, b=True)
        sheet['A5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet['A5'].border = thick_border

        sheet["B5"] = f"{survey_data['end_chainage']}"
        sheet['B5'].font = Font(size=13)
        sheet['B5'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet['B5'].border = thick_border

        img = Image(ROADATHENA_LOGO_PATH)
        img.anchor = 'A1'
        sheet.add_image(img)

        img = Image(HANUAI_LOGO_PATH)
        img.anchor = 'K1'
        sheet.add_image(img)

        sheet.row_dimensions[1].height = 60
        sheet.row_dimensions[2].height = 30
        sheet.row_dimensions[3].height = 30
        sheet.row_dimensions[4].height = 50
        sheet.row_dimensions[5].height = 50
        sheet.row_dimensions[6].height = 50

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20

        return sheet

    def create_detailed_report(data, survey_data, output_folder):

        #check if anomalies are enabled
        excel_path= os.path.join(output_folder, f"{roadId}_formatted.xlsx")
        if os.path.isfile(excel_path):
            #anomalies enabled, open current workbook
            workbook = load_workbook(os.path.join(output_folder, f"{roadId}_formatted.xlsx"))
            sheet_assets = workbook.create_sheet(title="Assets")
        else:
            #anomalies disabled, create new workbook
            workbook= Workbook()
            sheet_assets = workbook.active
            sheet_assets.title = "Assets"

        # Apply common formatting (assuming this function exists)
        sheet_assets = apply_common_formatting(sheet_assets, survey_data)

        # Define border style
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Define headers
        headers = [
            "Assets Number", "Timestamp on processed video", "Asset type", "Side", 
            "Category", "Latitude", "Longitude", "Distance from start point in meters", "Image Link"
        ]
        
        for col_num, header in enumerate(headers, start=1):
            cell = sheet_assets.cell(row=6, column=col_num, value=header)
            cell.font = Font(b=True, size=15)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = PatternFill(start_color='c0d3eb', end_color='c0d3eb', patternType='solid')

        serial_number = 1
        row_index = 8  # Start from row 8


        sorted_assets = sorted(
        data["assets"],
        key=lambda x: float(x.get("Distance", float("inf")))
        )

        for row_item in sorted_assets:
            asset_type = row_item.get("Asset type")
            if asset_type in ["NON_STANDARD_INFOMATORY_SIGNS", "ADVERTISEMENT_ENCHROACHMENT_SIGNS", "FADED_INFORMATORY_SIGNS", "DAMAGED_SIGN"]:
                continue

            lat, lng = row_item.get("Latitude"), row_item.get("Longitude")
            if lat in (0, None) or lng in (0, None):
                continue

            sheet_assets[f'A{row_index}'] = serial_number
            serial_number += 1
            sheet_assets[f'B{row_index}'] = row_item.get("Timestamp on processed video", "")
            sheet_assets[f'C{row_index}'] = asset_type
            sheet_assets[f'D{row_index}'] = row_item.get("Side", "")
            sheet_assets[f'E{row_index}'] = row_item.get("category", "")
            sheet_assets[f'F{row_index}'] = lat
            sheet_assets[f'G{row_index}'] = lng
            sheet_assets[f'H{row_index}'] = row_item.get("Distance", "")
            sheet_assets[f'I{row_index}'] = row_item.get("image", "")

            for col in range(1, 10):
                cell = sheet_assets.cell(row=row_index, column=col)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.border = thin_border

            row_index += 1 

        road_data = fetch_road_data(road_id)
        output_file_path = os.path.join(output_folder, f"{road_id}_formatted.xlsx")
        workbook.save(output_file_path)


    # Load JSON data and fetch road data
    file_path = os.path.join(output_json_path, f"road_{roadId}.json")

    with open(file_path, "r") as file:
        json_data = json.load(file)

    road_data = fetch_road_data(road_id)
    
    if road_data:
        try:
            seg = road_data["road"]["name"]
            start_chainage = road_data["road"]["start_chainage"]
            end_chainage = road_data["road"]["end_chainage"]
            mc_name = road_data["road"]["assigned_to"]["username"]
            sub_division_name = road_data["road"]["assigned_to"]["sub_division"]
            key = road_data["created_at"].split("T")[0]
        except KeyError as e:
            print(f"KeyError: {e}")
            return

        survey_data = {
            "road_name": seg,
            "start_chainage": start_chainage,
            "end_chainage": end_chainage,
            "mc": {"name": mc_name},
            "sub_division": {"sub_division": sub_division_name},
            "survey_date": key
        }

 
        create_detailed_report(json_data, survey_data, output_folder)
    else:
        print("Failed to fetch road data. Exiting...")


